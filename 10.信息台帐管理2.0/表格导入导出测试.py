from openpyxl import load_workbook

# 读取表格文件保存到wb对象
wb = load_workbook('backup.xlsx')
# 获取子表名称（表单） sheetnames 属性名 类型是列表
# print(wb.sheetnames)
# print(wb.get_sheet_names()) #返回列表
# 另外一种方法来获取表单名称

print(wb.sheetnames[0],end='\n\n')
# for sheet in wb:
#     print(sheet.title)
# 表单1对象生成（banding代表子表1）
banding = wb['绑定表']
print(banding)
# 活动子表默认就是第一个，所以可以这样
# banding = wb.active
# print(banding)
print(banding['a1'].value)
# print(banding.cell(row = 2,column = 2).value)
# 已知长宽遍历表格10*4 第一行是标题不管
for i in range(2,11):
    for j in range(1,5):
        t = banding.cell(row = i,column = j).value
        print(t,end='\t')
    print('当前行结束')

# for cell in banding:





wb.close()