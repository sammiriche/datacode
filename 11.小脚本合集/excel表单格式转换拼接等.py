# 通用模板，主要用作对excel特定表项做格式转化处理

from openpyxl import load_workbook, Workbook


class Excel_handle_model(object):

    # 1，转化360里面的mac格式为交换机识别的格式
    def convert_mac(self, path):
        self.wb = load_workbook(path)  # 实际路径里面的斜杠需要转义
        self.ws = self.wb.active  # 注意后面没有括号。使用了修饰器转换方法为属性
        # 获取行数
        row_num = self.ws.max_row
        # 假设mac地址在第三列,vlan在第四列
        # 使用for循环读取每一行的数据，并且进行转换处理
        for i in range(2, row_num + 1):  # 一般第一行为抬头，不用处理，根据距离情况修改
            content = self.ws.cell(row=i, column=3).value  # 取值str类型
            vlan = self.ws.cell(row=i, column=4).value  # 取得vlan值并且做合成之用
            content = content.strip()  # 去除收尾空格
            content = content.replace(':', "")  # 去除冒号
            ls_content = list(content)
            ls_content.insert(0, 'mac-address blackhole  ')
            ls_content.insert(5, '-')
            ls_content.insert(10, '-')
            ls_content.append(' ' + str(vlan))  # 末尾添加，不需要下标
            content = ''.join(ls_content)  # 列表转字符串
            self.ws.cell(row=i, column=3).value = content  # 重新赋值给表格

            # 输出到txt文件
            with open('g:\\333.txt', 'a+') as f:
                f.writelines(content)
                f.writelines('\n')

        # 可以不用Workbook新建表格对象
        self.wb.save('g:\\222.xlsx')


if __name__ == "__main__":
    ehm = Excel_handle_model()
    ehm.convert_mac('g:\\111.xlsx')