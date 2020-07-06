'''
基本流程
1：设计基本UI，转换成py文本
2：从表格获取域名函数，更新hosts到本地函数
3：信号槽绑定

'''
from openpyxl import load_workbook
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QDialog,QApplication,QFileDialog,QMessageBox

import sys

class Update_hosts(QDialog):
    def __init__(self):
        super().__init__()
        # 最终需要导入的域名（包含格式）
        format_add = ''
        self.setupUi(self)

    def setupUi(self, Dialog):
        Dialog.setObjectName("Dialog")
        Dialog.resize(257, 493)
        self.add_list = QtWidgets.QPlainTextEdit(Dialog)
        self.add_list.setGeometry(QtCore.QRect(20, 70, 211, 351))
        font = QtGui.QFont()
        font.setPointSize(9)
        self.add_list.setFont(font)
        self.add_list.setPlainText("")
        self.add_list.setObjectName("add_list")
        self.titel_label = QtWidgets.QLabel(Dialog)
        self.titel_label.setGeometry(QtCore.QRect(50, 30, 141, 21))
        font = QtGui.QFont()
        font.setFamily("微软雅黑")
        font.setPointSize(16)
        self.titel_label.setFont(font)
        self.titel_label.setObjectName("titel_label")
        self.import_btn = QtWidgets.QPushButton(Dialog)
        self.import_btn.setGeometry(QtCore.QRect(20, 430, 101, 23))
        self.import_btn.setObjectName("import_btn")
        self.update_btn = QtWidgets.QPushButton(Dialog)
        self.update_btn.setGeometry(QtCore.QRect(128, 430, 101, 23))
        self.update_btn.setObjectName("update_btn")

        self.retranslateUi(Dialog)
        QtCore.QMetaObject.connectSlotsByName(Dialog)

    def retranslateUi(self, Dialog):
        _translate = QtCore.QCoreApplication.translate
        Dialog.setWindowTitle(_translate("Dialog", "Dialog"))
        self.titel_label.setText(_translate("Dialog", "更新HOST文件"))
        self.import_btn.setText(_translate("Dialog", "从表格导入域名"))
        self.update_btn.setText(_translate("Dialog", "更新域名到hosts"))

        # 信号槽绑定
        self.import_btn.clicked.connect(self.import_clicked)
        self.update_btn.clicked.connect(self.update_clicked)

    def update_clicked(self):
        self.write_to_hosts()
    def import_clicked(self):
        # 注意这里返回的路径是列表文件
        # 导入之前清空文本框
        self.add_list.setPlainText('')
        path = QFileDialog.getOpenFileName(self,'请选择需要导入的文件','C:\\')
        self.get_add(path[0])
        for i in self.add:
            # 这里用append可以循环追加。直接用settext会覆盖旧内容
            self.add_list.appendPlainText(i)
        # 获取文本内容方法和lineedit名称不一样    
        print(self.add_list.toPlainText())
        self.format_add = self.add_list.toPlainText()

    def get_add(self,path):
        # 从表格获取恶意域名并且保存到列表,注意不要重名
        self.add = []
        wb = load_workbook(path)
        ws = wb.active
        # 获取表单行数
        row_num = ws.max_row
        for i in range(2,row_num+1):
            ip = ws.cell(row = i,column = 3).value
            ip = 'localhost    ' + ip
            self.add.append(ip)

    def write_to_hosts(self):
        self.format_add = self.add_list.toPlainText()
        if self.format_add == '':
            print('内容为空')
            reply = QMessageBox.about(self,'提示','请导入域名列表或者手动添加')
        else:
            # 以追加模式打开本地的hosts文件
            with open(r'C:\Windows\System32\drivers\etc\hosts','a+',encoding='UTF-8') as file:
                file.write('\n')
                file.write(self.format_add)
                reply = QMessageBox.about(self,'提示','本地hosts文件更新成功')
                # 更新成功后清空文本框
                self.add_list.setPlainText('')

if __name__ == "__main__":
    app = QApplication(sys.argv)
    uh = Update_hosts()
    uh.show()
    sys.exit(app.exec_())
