import re
import pymysql
import openpyxl
import os


# 主程序文件
class Ems(object):
    def __init__(self):
        # 创建数据库和表
        conn = pymysql.connect(host='localhost',
                               user='root',
                               passwd='root',
                               port=3306,
                               charset='utf8mb4')
        cur = conn.cursor()
        # 创建数据库默认为utf8mb4字符集
        sql = 'create database if not exists ems DEFAULT CHARSET utf8mb4 COLLATE utf8mb4_unicode_ci'
        cur.execute(sql)
        cur.execute('use ems')  # 使用当前数据库，
        # 创建表
        sql = '''
            create table if not exists employees(
                em_name varchar(20),
                em_dept varchar(20),
                em_ip varchar(20),
                em_mac varchar(20)
            )
        '''
        cur.execute(sql)
        cur.close()
        conn.close()

        print('欢迎使用员工管理系统')
        print('*' * 62)
        # 进入主函数
        self.run()

    # 检查输入合法性函数系列
    # 检查数字输入合法性
    def is_num(self):
        while 1:
            num = input('请输入1-8之间的正整数：')
            pattern = r'^[1-8]$'
            result = re.match(pattern, num)
            if result:
                return int(num)
            else:
                print('请检查输入格式')

    # 检查名字输入合法性
    def is_name(self):
        while 1:
            name = input('请输入姓名：')
            pattern = r'^[\u4e00-\u9fa5]{2,4}$'
            result = re.match(pattern, name)
            if result:
                return name
            else:
                print('请检查输入格式')

    def is_dept(self):
        while 1:
            dept = input('请输入部门：')
            pattern = r'^[\u4e00-\u9fa5]{2,6}$'
            result = re.match(pattern, dept)
            if result:
                return dept
            else:
                print('请检查输入格式')

    def is_ip(self):
        while 1:
            ip = input('请输入IP地址：')
            pattern = r'^((25[0-4]|2[0-4]\d|1\d\d|\d?\d)\.){3}(25[0-4]|2[0-4]\d|1\d\d|\d?\d)$'
            result = re.match(pattern, ip)
            if result:
                return ip
            else:
                print('请检查输入格式')

    def is_mac(self):
        while 1:
            mac = input('请输入mac地址：')
            pattern1 = r'^(([0-9a-fA-F]{4}\.){2}[0-9a-fA-F]{4})$'  # xxxx.xxxx.xxxx
            pattern2 = r'^(([0-9a-fA-F]{4}-){2}[0-9a-fA-F]{4})$'  # xxxx-xxxx-xxxx
            pattern3 = r'^(([0-9a-fA-F]{2}:){5}[0-9a-fA-F]{2})$'  # xx:xx:xx:xx:xx:xx
            result1 = re.match(pattern1, mac)
            result2 = re.match(pattern2, mac)
            result3 = re.match(pattern3, mac)
            if result1 or result2 or result3:
                mac = self.trans_mac(mac)
                return mac
            else:
                print('请检查输入格式')

    # 统一mac格式函数
    def trans_mac(self, mac):
        if '-' in mac:
            return mac
        elif '.' in mac:
            mac = mac.replace('.', '-')
            return mac
        else:
            # 字符串删除指定位置字符转换为列表最简单 01:23:45:67:89:10:11
            mac = mac.replace(':', '')  # 去冒号
            ls_mac = list(mac)
            ls_mac.insert(4, '-')
            ls_mac.insert(9, '-')
            # 从列表转换为字符串
            mac = ''.join(ls_mac)
            return mac

    # 数据库操作函数
    # 数据库连接
    def con_sql(self):
        conn = pymysql.connect(host='localhost',
                               user='root',
                               passwd='root',
                               port=3306)
        cur = conn.cursor()
        return conn, cur

    def close_sql(self, conn, cur):
        conn.commit()
        cur.close()
        conn.close()

    # 增删改查等功能函数
    def add_em(self):
        name = self.is_name()
        dept = self.is_dept()
        ip = self.is_ip()
        mac = self.is_mac()
        # 从返回值获取连接和游标
        conn, cur = self.con_sql()
        cur.execute('use ems')
        sql_insert = 'insert into employees values(%s,%s,%s,%s)'
        cur.execute(sql_insert, (name, dept, ip, mac))
        self.close_sql(conn, cur)

    def del_em(self):
        print('根据姓名删除，请按提示操作！')
        name = self.is_name()
        conn, cur = self.con_sql()
        cur.execute('use ems')
        sql_del = 'delete from employees where em_name = %s'
        cur.execute(sql_del, name)  # 后面参数也可以写成(name,)
        self.close_sql(conn, cur)

    def modify_em(self):
        name = self.is_name()
        conn, cur = self.con_sql()
        cur.execute('use ems')
        #  先判断是否存在
        sql = 'select * from employees where em_name = %s'
        cur.execute(sql, name)
        if cur.rowcount:  # 代表结果有数据
            dept = self.is_dept()
            ip = self.is_ip()
            mac = self.is_mac()
            sql_modify = 'update employees set em_dept = %s,em_ip = %s,em_mac = %s where em_name = %s'
            cur.execute(sql_modify, (dept, ip, mac, name))
            print('信息修改成功')
        else:
            print('数据库没有当前此人信息')
        self.close_sql(conn, cur)

    def query_em(self):
        print('按1，根据姓名查询 | 按2，根据IP地址查询')
        num = int(input('press 1 or 2 \n'))
        if num == 1:
            conn, cur = self.con_sql()
            name = self.is_name()
            cur.execute('use ems')
            sql = 'select * from employees where em_name = %s'
            cur.execute(sql, name)
            if cur.rowcount > 1:
                for i in range(0, cur.rowcount):
                    result = cur.fetchone()
                    print(result)
            elif cur.rowcount == 1:
                result = cur.fetchone()
                print(result)
            else:
                print('没有当前员工信息')
            self.close_sql(conn, cur)  # 如果直接按了2.那么这里的conn就会没有定义了

        if num == 2:
            conn, cur = self.con_sql()
            ip = self.is_ip()
            cur.execute('use ems')
            sql = 'select * from employees where em_ip = %s'
            cur.execute(sql, ip)
            if cur.rowcount > 1:
                for i in cur.fetchall():
                    print(i)
            elif cur.rowcount == 1:
                print(cur.fetchone())
            else:
                print('没有当前关联IP信息')
            self.close_sql(conn, cur)

    def show_em(self):
        conn, cur = self.con_sql()
        cur.execute('use ems')
        sql_show = 'select * from employees'
        cur.execute(sql_show)
        result = cur.fetchall()

        # 先打印表头
        print('姓名'.ljust(8), end='\t')
        print('部门'.ljust(8), end='\t')
        print('IP'.ljust(8), end='\t')
        print('MAC'.ljust(8))
        # result是由元组构成的元组 注意循环的意思
        for i in result:
            for j in i:
                print(j.ljust(8), end='\t')
            print()
        # print(result)
        self.close_sql(conn, cur)

    def import_excel(self):
        flag = input('请按任意键导入表格，按n取消 \n')
        if not flag == 'n':
            # 先连接数据库，方便后面对接
            conn, cur = self.con_sql()
            cur.execute('use ems')

            # 先判断是否存在模板表格
            if os.path.exists('template.xlsx'):
                # 获取表格对象
                wb = openpyxl.load_workbook('template.xlsx')
                # 获取表单对象
                ws = wb.active  # 后面没有括号
                # 遍历单元格读取其中的值,首先获取行数和列数
                row_num = ws.max_row  #行数，注意第一行为抬头，不用读取
                for i in range(2, row_num + 1):
                    name = ws.cell(row=i, column=1).value
                    dept = ws.cell(row=i, column=2).value
                    ip = ws.cell(row=i, column=3).value
                    mac = ws.cell(row=i, column=4).value
                    # 信息保存进数据库
                    sql = 'insert into employees values(%s,%s,%s,%s)'
                    cur.execute(sql, (name, dept, ip, mac))
                self.close_sql(conn, cur)
                print('表格导入成功')
        else:
            print('您已经取消当前操作！')

    def export_excel(self):
        # 首先打开数据库。读取所有数据库保存到元组或者列表？
        conn, cur = self.con_sql()
        cur.execute('use ems')
        sql = 'select * from employees'
        cur.execute(sql)
        result = cur.fetchall()  # 二维元组

        # 打开表格，定位到表单
        wb = openpyxl.Workbook()  # 创建新表
        ws = wb.active
        ws.title = '绑定表'
        ws['a1'] = '姓名'
        ws['b1'] = '部门'
        ws['c1'] = 'IP'
        ws['d1'] = 'MAC'

        # for循环读取每一个元素赋值到对应对象
        num = 2  # 控制行标作用
        for i in result:
            # 将对象保存到表格
            ws.cell(row=num, column=1).value = i[0]
            ws.cell(row=num, column=2).value = i[1]
            ws.cell(row=num, column=3).value = i[2]
            ws.cell(row=num, column=4).value = i[3]
            num += 1
        # 保存表格
        wb.save('backup.xlsx')
        # 关闭数据库
        self.close_sql(conn, cur)
        print('数据导出成功。请用excel查看！')

    def run(self):
        # 显示欢迎界面和选择菜单，因为需要一直显示，使用while循环
        while True:
            print('*' * 62)
            print('1:添加员工信息'.ljust(8), end='\t')
            print('2:删除员工信息'.ljust(8), end='\t')
            print('3:修改员工信息'.ljust(8), end='\t')
            print('4:查询员工信息')
            print('5:显示所有员工'.ljust(8), end='\t')
            print('6:表格批量导入'.ljust(8), end='\t')
            print('7:表格批量导出'.ljust(8), end='\t')
            print('8:退出当前系统')
            print('*' * 62)

            # 进入功能菜单
            num = self.is_num()
            if num == 1:
                self.add_em()
            if num == 2:
                self.del_em()
            if num == 3:
                self.modify_em()
            if num == 4:
                self.query_em()
            if num == 5:
                self.show_em()
            if num == 6:
                self.import_excel()
            if num == 7:
                self.export_excel()
            if num == 8:
                break


if __name__ == '__main__':
    ems = Ems()
