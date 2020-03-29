import openpyxl
# 获取表格对象,注意文件路径是否生效，测试
wb = openpyxl.load_workbook('template.xlsx')

# getting sheets name from the workbook
name = wb.sheetnames
# print(name)
# print(type(name))
# getting sheet object from the workbook
ws = wb.active
# ws = wb['绑定表']
print(ws)
#getting cells from the sheet
print(ws['a2'].value)
v = ws['b3']
print(v.coordinate,v.value) # coordinate相当于单元格坐标

result = ws.cell(row = 3,column = 2).value
print(result)
# 循环打印表格所有对象
for i in range(0,5):
    for j in range(0,4):
        result = ws.cell(row = i+1,column = j+1).value
        print(result,end = '\t')
    print('')