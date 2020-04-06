# 在表格管理类主要实现数据库和表格的联动。实现相互之间的格式转换
from openpyxl import load_workbook,Workbook
from Mysql_manager import *
import os

class Excel_manager(object):
    # def __init__(self):
    #     print('请确认是否进行表格的导入导出操作y or n')
    #     self.flag = input('') # 该标签在其他函数调用。所以保存到类属性


    def import_excel(self):
        # 首先判断需要导入的表格是否存在。防止后面报错
        if os.path.exists('template.xlsx'):
            # 首先获取表格对象
            # 一次添加，最后commit
            mm = Mysql_manager('localhost','root','root',3306,'milkbottle')
            with mm:
                sql = 'insert into em_info values(%s,%s,%s,%s)'
                wb = load_workbook('template.xlsx')
                # 然后获取工作单子表
                ws = wb.active
                # 获取工作单行数
                row_num = ws.max_row # 属性，不是函数max_row  第一行为表头
                for i in range(2,row_num+1): # 表格是从1开始，而不是0
                    name = ws.cell(row = i,column = 1).value
                    dept = ws.cell(row = i,column = 2).value
                    ip = ws.cell(row = i,column = 3).value
                    mac = ws.cell(row = i,column = 4).value
                    mm.cur.execute(sql,(name,dept,ip,mac))
            print('从表格导入信息成功')
        else:
            print('请在指定目录存放template.xlsx文件')

    def export_excel(self):
        # 首先连接数据库，并且读出所有信息
        mm = Mysql_manager('localhost','root','root',3306,'milkbottle')
        # 在内存中创建一个表格
        wb = Workbook()
        ws = wb.active
        # 输出表头
        ws.cell(row = 1,column = 1).value = '姓名'
        ws.cell(row = 1,column = 2).value = '部门'
        ws.cell(row = 1,column = 3).value = 'IP地址'
        ws.cell(row = 1,column = 4).value = 'MAC地址'
        with mm:
            sql = 'select * from em_info'
            mm.cur.execute(sql)
            result = mm.cur.fetchall() # 取出所有信息保存在二维元组里面
            num = 2
            for i in result:
                # 这里的i就是每一个单独员工的所有信息
                # 注意从表格的第二行开始写入
                ws.cell(row = num,column = 1).value = i[0]
                ws.cell(row = num,column = 2).value = i[1]
                ws.cell(row = num,column = 3).value = i[2]
                ws.cell(row = num,column = 4).value = i[3]
                num += 1
            # 最后需要保存表格，不然都是在内存操作
            wb.save('backup.xlsx')
            print('信息成功导出到表格')

if __name__ == '__main__':
    ex = Excel_manager()
    ex.import_excel()


