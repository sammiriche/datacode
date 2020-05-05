# 用于ui界面的导入和导出操作封装 excel和数据库之间数据传递

from openpyxl import load_workbook,Workbook
from Mysql_manager import *

class Excel_manager(object):
    # 从表格导入到数据库
    def import_excel(self,path):
        # 表格对象
        self.wb = load_workbook(path)
        # 表单（子表）对象
        self.ws = self.wb.active
        
        # 建立数据库连接
        mm = Mysql_manager()
        sql = 'insert into em_info values(%s,%s,%s,%s)' 
        with mm:
            # 循环读出，循环写入
            row_num = self.ws.max_row # 获取子表行数
            print('测试')
            print(row_num)
            # 注意表格计数是从1开始，不是0
            for i in range(2,row_num+1):
                name = self.ws.cell(row = i,column = 1).value
                dept = self.ws.cell(row = i,column = 2).value
                ip = self.ws.cell(row = i,column = 3).value
                mac = self.ws.cell(row = i,column = 4).value

                mm.db_exe(sql,(name,dept,ip,mac))
                
                print('数据成功导入到数据库')

    def export_excel(self,path):
        self.wb = Workbook()
        self.ws = self.wb.active
        # 先输出表头
        self.ws.cell(row = 1,column = 1).value = '姓名'
        self.ws.cell(row = 1,column = 2).value = '部门'
        self.ws.cell(row = 1,column = 3).value = 'IP地址'
        self.ws.cell(row = 1,column = 4).value = 'MAC地址'
        mm = Mysql_manager()
        with mm:
            result = mm.db_show()
            num = 2 # 计数递增使用
            for i in result:
                self.ws.cell(row = num,column = 1).value = i[0]
                self.ws.cell(row = num,column = 2).value = i[1]
                self.ws.cell(row = num,column = 3).value = i[2]
                self.ws.cell(row = num,column = 4).value = i[3]

                num += 1
            #最后记得需要保存
        print(f'数据成功导出到{path}')
        self.wb.save(path)
             

