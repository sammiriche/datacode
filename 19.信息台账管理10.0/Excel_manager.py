# 数据库和电子表格格式的相互转换
from openpyxl import load_workbook,Workbook
from Mysql_manager import *

class Excel_manager(object):
    def import_excel(self,path):
        # path 通过其他方法获取并传递进来
        # 从表格导入数据到数据库
        # 先从打开表格,获取当前子表
        self.wb = load_workbook(path)
        self.ws = self.wb.active
        # 使用for循环读出相应数据
        # 首先获取表行数
        row_num = self.ws.max_row #都是属性，不是方法 使用了修饰器的方法
        mm = Mysql_manager()
        with mm:
            sql = 'insert into em_info values(%s,%s,%s,%s,%s,%s,%s)'
            for i in range(2,row_num+1):   # 表格是从1开始计数，去掉抬头
                name = self.ws.cell(row = i,column = 1).value
                dept = self.ws.cell(row = i,column = 2).value
                ip = self.ws.cell(row = i,column = 3).value
                mac = self.ws.cell(row = i,column = 4).value
                room = self.ws.cell(row = i,column = 5).value
                switch = self.ws.cell(row = i,column = 6).value
                port = self.ws.cell(row = i,column = 7).value
                mm.exe_db(sql,(name,dept,ip,mac,room,switch,port))
            print('数据导入成功')
            
    def export_excel(self,path):
        # 数据导出为表格
        # 先实例表格对象
        self.wb = Workbook()
        self.ws = self.wb.active
        # 先写入表头
        self.ws.cell(row = 1,column = 1).value = '姓名'
        self.ws.cell(row = 1,column = 2).value = '部门'
        self.ws.cell(row = 1,column = 3).value = 'IP地址'
        self.ws.cell(row = 1,column = 4).value = 'MAC地址'
        self.ws.cell(row = 1,column = 5).value = '房间号'
        self.ws.cell(row = 1,column = 6).value = '交换机'
        self.ws.cell(row = 1,column = 7).value = '端口'
        # 从数据库读取数据
        mm = Mysql_manager()
        with mm:
            result = mm.show_db() # 返回二元数组
            num = 2 # 从第二行开始写入
            for i in result:
                self.ws.cell(row = num,column = 1).value = i[0]
                self.ws.cell(row = num,column = 2).value = i[1]
                self.ws.cell(row = num,column = 3).value = i[2]
                self.ws.cell(row = num,column = 4).value = i[3]
                self.ws.cell(row = num,column = 5).value = i[4]
                self.ws.cell(row = num,column = 6).value = i[5]
                self.ws.cell(row = num,column = 7).value = i[6]
                num +=1
        self.wb.save(path)
        print('数据导出成功')