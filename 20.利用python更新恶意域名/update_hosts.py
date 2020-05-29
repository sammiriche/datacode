'''
基本流程
1：打开本身hosts文件
复习文件写入的几种方式
2：利用for循环读取已知excel文件的内容，并且循环写入到文件当中
   循环体字符串变量拼接localhost    www.sina.com.cn
3：保存


'''
# openpyxl 很多都是属性形式
from openpyxl import load_workbook,Workbook

class Update_hosts(object):
    # 文件操作函数
    def operator_file(self):
        pass
    def load_excel(self,path):
        # 获取表格文件
        wb = load_workbook(path)
        # 获取工作单
        ws = wb.active
        # 获取工作单行数
        row_num = ws.max_row
        