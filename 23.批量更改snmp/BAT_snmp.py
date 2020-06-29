from openpyxl import load_workbook
import paramiko

# 从表格导入交换机地址，并且进行版本，品牌分类
class Switch_statis(object):
    def __init__(self):
        # 建立不同版本的IP列表，方便后期遍历
        self.huawei_ls_1 = []
        self.huawei_ls_2 = []
        self.h3c_ls_v3 = []  # 早期H3C交换机软件版本comware v3
        self.h3c_ls_v5 = []  # 最普遍版本
        self.h3c_ls_v7 = []  # 最新版本，比如3100V3。命令有较大变动

    # 从统计表格获取所有地址然后分类 
    def collect_ip(self):
        wb = load_workbook(path)
        ws = wb.active
        row_num = ws.max_row
        # 假设IP地址在第二列
        for i in range(2,row_num):
            ip = ws.cell(row = i,column = 2).value
            # 实例化SSH连接
            ssh = paramiko.SSHClient()
            # 自动添加进信任列表
            ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            ssh.connect(
                host = ip,
                port = 22,
                name = 'admin',
                password = 'xtgsyjb.6193',
                allow_agent=False,
                look_for_keys=False
            )
            stdin,stdout,stderr = ssh.exec_command('display version')
            result = stdout.read().decode()

            # 查找子字符串进行类型判断
            if 'H3C' in result:
                print(f'{ip}交换机型号为H3C')
                # 对H3C进行软件版本分类
                # 
            else:
                print(f'{ip}交换机型号为华为')
