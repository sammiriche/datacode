# 控制面板主模块，相关功能主要在这里实现
import re
import pickle
import os
from empolyee2 import Empolyee
from exceldata import Excel
from openpyxl import load_workbook
from openpyxl import Workbook


class Control(object):
    def __init__(self):
        # 列表用于保存所有学生信息
        self.emList = []
        # 如果data2是空，那么load会报错
        if os.path.getsize('data2'):
            with open('data2', 'rb') as f:
                self.emList = pickle.load(f)

    def run(self):
        # 这里的双重循环是为了配合子函数，比如添加。里面正则的判断完成之后。重新从13行开始
        flag = 1
        while flag:
            while True:
                self.show_info()
                num = input('请输入您要进行的操作编号：')
                # 判断输入数值合法性
                bl = re.match(r'^[1-8]$', num)
                # bl = re.match(r'[1-9]([0-9]{4,10})',num)
                # print(bl)
                if not bl:
                    print('请输入1到8的整数！')
                else:
                    break
            num = int(num)
            if num == 1:
                self.emShow()
            if num == 2:
                self.emSearch()
            if num == 3:
                self.emAdd()
            if num == 4:
                self.emDel()
            if num == 5:
                self.emModify()
            if num == 6:
                self.emSave()
            if num == 7:
                self.emBatch()
            if num == 8:
                print('您已退出当前系统！')
                flag = 0
                break

    # 定义静态方法，加了修饰器，那么括号里面不用加self，不调用任何属性
    @staticmethod
    def show_info():
        print('*' * 25)
        print('欢迎使用员工台帐2.0')
        print('*' * 25)
        print('1：显示所有员工信息')
        print('2：查询指定员工信息')
        print('3：添加员工信息')
        print('4：删除员工信息')
        print('5：修改员工员工信息')
        print('6：保存员工信息')
        print('7：批量操作员工信息')
        print('8：退出当前系统')
        print('*' * 25)

    def emShow(self):
        # 注意对齐问题。字符串补全 左对齐。不足补到8.字符串变量最低长度为8了
        flag = 0
        while not flag:
            if len(self.emList) == 0:
                print('没有任何员工信息')
                break
            else:
                print('姓名'.ljust(8), end='\t')
                print('部门'.ljust(8), end='\t')
                print('IP'.ljust(8), end='\t')
                print('MAC'.ljust(8))
                for i in self.emList:
                    print(i.getName().ljust(8), end='\t')
                    print(i.getDept().ljust(8), end='\t')
                    print(i.getIp().ljust(8), end='\t')
                    print(i.getMac().ljust(8))
                flag = input('按任意键回车返回主菜单：')

    def emSearch(self):
        pass

    # 里面加上while和if语句方便输入检测。替代goto语句
    def emAdd(self):
        while True:
            name = input('请输入员工姓名：')
            # 这里正则必须假如前后的两个特殊符号，不然3位以上也能匹配
            bln = re.match(r'^[\u4e00-\u9fa5]{2,4}$', name)
            if not bln:
                print('请输入正确的姓名（1-3位中文字符）')
            else:
                break
        # 关于部门选择可以考虑做成列表选择
        while True:
            dept = input('请输入员工部门：')
            bld = re.match(r'^[\u4e00-\u9fa5]{2,6}$', dept)
            if not bld:
                print('请输入正确的部门')
            else:
                break
        while True:
            ip = input('请输入员工IP：')
            # ip地址的正则匹配。可以带前导0
            bli = re.match(
                r'^((25[0-4]|2[0-4]\d|1\d\d|\d?\d)\.){3}(25[0-4]|2[0-4]\d|1\d\d|\d?\d)$',
                ip)
            if not bli:
                print('请输入点分十进制的ip地址')
            else:
                break
        while True:
            mac = input('请输入员工MAC：')
            p1 = re.compile(
                r'^([0-9a-fA-F]{4}\.){2}[0-9a-fA-F]{4}$')  # XXXX.XXXX.XXXX
            p2 = re.compile(
                r'^([0-9a-fA-F]{4}-){2}[0-9a-fA-F]{4}$')  # XXXX-XXXX-XXXX
            p3 = re.compile(
                r'^([0-9a-fA-F]{2}:){5}[0-9a-fA-F]{2}$')  # XX:XX:XX:XX:XX:XX
            blm1 = re.match(p1, mac)
            blm2 = re.match(p2, mac)
            blm3 = re.match(p3, mac)
            if blm1 or blm2 or blm3:
                break
            else:
                print('请输入正确格式的mac地址')
                print('格式为XXXX.XXXX.XXXX或者XXXX-XXXX-XXXX或者XX:XX:XX:XX:XX:XX')
        em = Empolyee(name, dept, ip, mac)
        self.emList.append(em)
        print('员工信息添加成功', end=' ')
        print(em)

        # 测试代码
        # for em in self.emList:
        #     print(em)

    def emDel(self):
        print('--按 1 选择根据姓名删除信息 | 按 2 选择根据IP删除信息--')
        num = int(input("press 1 or 2 "))
        if num == 1:
            name = input('请输入准备删除的员工姓名：')
            if len(self.emList) != 0:
                for i in self.emList:
                    if i.getName() == name:
                        self.emList.remove(i)
                        print(f'{i.getName()}的个人信息删除成功')
                    else:
                        print('该员工不存在!')
            else:
                print('当前没有任何员工信息')
        if num == 2:
            ip = input('请输入需要删除的ip地址：')
            if len(self.emList) != 0:
                for i in self.emList:
                    if i.getIp() == ip:
                        self.emList.remove(i)
                        print(f'{i.getName()}的个人信息删除成功')
                    else:
                        print('该IP地址信息不存在')
            else:
                print('当前没有任何IP信息')

    def emModify(self):
        while True:
            name = input('请输入准备修改的员工姓名：')
            pattern = re.compile(r'^[\u4e00-\u9fa5]{2,4}$')
            bl = re.match(pattern, name)
            if not bl:
                print('请输入正确的姓名格式：')
            else:
                break
        if len(self.emList) == 0:
            print('当前员工列表为空')
        else:
            for i in self.emList:
                if i.getName() == name:
                    i.setName(name)
                    i.setDept(input('请输入员工部门：'))
                    i.setIp(input('请输入员工IP：'))
                    i.setMac(input('请输入员工MAC：'))
                    print('员工信息修改成功')
                    print(i)

    def emSave(self):  # 使用with open 语句 在缩进完成后自动关闭文件
        # print(len(self.emList))
        # print(type(self.emList))
        with open('data2', 'wb') as f:
            pickle.dump(self.emList, f)
            print('信息更新成功')

    def emBatch(self):
        num = int(input('按1选择批量导入员工信息 | 按2选择批量导出员工信息: '))
        if num == 1:
            newExcel = Excel()
            # 调用表格类的导入方法
            newExcel.imExcel()
            # 类里的列表与表格里的列表合并操作
            self.emList += newExcel.lsExcel
        if num == 2:
            # 创建新表格文件
            wb = Workbook()
            # 默认就用个工作单，如果新建，就会是第二个。所以我们需要把第一个工作单改名
            # ws = wb.create_sheet('绑定表')
            ws  = wb.active
            ws.title = '绑定表'
            # 开始写入第一行信息
            ws['a1'] = '姓名'
            ws['b1'] = '部门'
            ws['c1'] = 'IP'
            ws['d1'] = 'MAC'
            rowNum = len(self.emList)
            i = 1
            for em in self.emList:
                # i控制表格行的变化
                ws.cell(row=i + 1, column=1).value = em.getName()
                ws.cell(row=i + 1, column=2).value = em.getDept()
                ws.cell(row=i + 1, column=3).value = em.getIp()
                ws.cell(row=i + 1, column=4).value = em.getMac()
                i += 1
            wb.save('backup.xlsx')


if __name__ == '__main__':
    cr = Control()
    cr.run()