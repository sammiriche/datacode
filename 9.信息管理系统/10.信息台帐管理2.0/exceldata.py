# 表格形式的数据导入导出工作
import os
from empolyee2 import Empolyee
from openpyxl import load_workbook
# from control_panel2 import Control
class Excel(object):
    def __init__(self):
        # 建立临时列表 后期和原始列表添加
        self.lsExcel = []
    # 表格导入
    def imExcel(self):
        if os.path.exists('backup.xlsx') and os.path.getsize(('backup.xlsx')):
            # 加载表格文件
            wb = load_workbook('backup.xlsx')
            # 获取工作单（子表）
            ws = wb.active
            # 获取最大行数
            rowNum = ws.max_row
            for i in range(2,rowNum+1):
                name = ws.cell(row = i,column = 1).value
                dept = ws.cell(row = i,column = 2).value
                ip = ws.cell(row = i,column = 3).value
                mac = ws.cell(row = i,column = 4).value
                em = Empolyee(name,dept,ip,mac)
                self.lsExcel.append(em)

# 做导出excel需要引用主面板的


