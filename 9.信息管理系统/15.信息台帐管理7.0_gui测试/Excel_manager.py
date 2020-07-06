# 数据库和电子表格数据的相互转换工作

from openpyxl import load_workbook,Workbook
import os
from Mysql_manager import *
class Excel_manager(object):

    def import_excel(self):
        # 导入数据库,先判断是否存在待导入文件
        if os.path.exists('template.xlsx'):
            # 建立表格对象
            wb = load_workbook('template.xlsx')
            # 获取表单对象
            ws = wb.active
            # 获取表格行数,注意后面没有括号
            row_num = ws.max_row
            # 实例数据库操作类。建立连接
            mm = Mysql_manager('localhost','root','root',3306,'milkbottle')
            with mm:
                sql = 'insert into em_info values(%s,%s,%s,%s)'
                # 表格是1开始计数。第二行开始是需要的数据
                for i in range(2,row_num+1):
                    name = ws.cell(row = i,column = 1).value
                    dept = ws.cell(row = i,column = 2).value
                    ip = ws.cell(row = i,column = 3).value
                    mac = ws.cell(row = i,column = 4).value
                    mm.cur.execute(sql,(name,dept,ip,mac))
                    print('数据导入成功')

        else:
            print('没有找到表格文件')

    def export_excel(self):
        wb = Workbook()
        ws = wb.active
        # 输出表头
        ws.cell(row = 1, column = 1).value = '姓名'
        ws.cell(row = 1, column = 2).value = '部门'
        ws.cell(row = 1, column = 3).value = 'ip'
        ws.cell(row = 1, column = 4).value = 'mac'
        mm = Mysql_manager('localhost','root','root',3306,'milkbottle')
        with mm:
            sql = 'select * from em_info'
            mm.cur.execute(sql)
            result = mm.cur.fetchall()
            num = 2 # 从第二行开始添加
            for i in result:
                # 每一个二维元组第一次拆包就是一个一元元组
                ws.cell(row = num,column = 1).value = i[0]
                ws.cell(row = num,column = 2).value = i[1]
                ws.cell(row = num,column = 3).value = i[2]
                ws.cell(row = num,column = 4).value = i[3]
                num +=1
            print('数据导出成功')
        wb.save('backup.xlsx')



