# 管理界面，也是一个类封装起来
# 先大的分解类，再分解功能函数，然后再细化。
import os
import re
from pa_empoloyee import Empo
from openpyxl import load_workbook
from openpyxl import Workbook


class Manager(object):
    # 生成对象时首先自动生成一个列表，用于保存所有员工信息
    # 并且加载保存好的员工信息（加载电子表格，并且提前判断）
    def __init__(self):
        # 导入表格信息，方法后面单独写
        self.em_list = []  # 这个变量必须在这里建立
        self.imExcel()
        self.run()

    def imExcel(self):
        if os.path.exists('em_data.xlsx'):
            #加载表格文件，赋值给wb对象
            wb = load_workbook('em_data.xlsx')
            # 获取wb的活动工作表，一般就是光标最后编辑所在，如果就一张表，就是当前表
            ws = wb.active
            # 从工作表读取信息。
            row_num = ws.max_row
            for i in range(2, row_num + 1):
                name = ws.cell(row=i, column=1).value
                dept = ws.cell(row=i, column=2).value
                ip = ws.cell(row=i, column=3).value
                mac = ws.cell(row=i, column=4).value
                em = Empo(name, dept, ip, mac)
                self.em_list.append(em)

    # 主程序窗口
    def run(self):
        flag = 1
        while True:
            if flag == 0:
                break
            print('*' * 40)
            print('1：显示所有员工信息'.ljust(12), end='\t')
            print('2：添加员工信息'.ljust(12))
            print('3：删除员工信息'.ljust(12), end='\t')
            print('4：修改员工信息'.ljust(12))
            print('5：查询员工信息'.ljust(12), end='\t')
            print('6：退出当前系统')
            print('*' * 40)

            # 注意不同地方的num的作用域 调用函数判断输入合法性，解决了多重循环问题
            num = self.isNum()
            if num == 1:
                self.show_em()
            if num == 2:
                self.add_em()
            if num == 3:
                self.del_em()
            if num == 4:
                self.modify_em()
            if num == 5:
                self.query_em()
            if num == 6:
                # 在退出前保存信息
                self.save_em()
                # 加标记，方便跳出外层循环
                flag = 0
                break

    # 各种功能函数定义
    def show_em(self):
        if len(self.em_list) == 0:
            print('当前没有任何员工信息')
        else:
            print('姓名'.ljust(8), end='\t')
            print('部门'.ljust(8), end='\t')
            print('IP'.ljust(8), end='\t')
            print('MAC'.ljust(8))
            for i in self.em_list:
                print(i.get_name().ljust(8), end='\t')
                print(i.get_dept().ljust(8), end='\t')
                print(i.get_ip().ljust(8), end='\t')
                print(i.get_mac().ljust(8))

    def add_em(self):
        name = self.isName()
        dept = self.isDept()
        ip = self.isIp()
        mac = self.isMac()
        em = Empo(name, dept, ip, mac)
        self.em_list.append(em)
        print(em)

    def del_em(self):
        # 因为有了下面判断语句，需要重新开始选择，所以需要加while语句
        while 1:
            num = int(input('按名字删除请按1 | 按IP删除请按2  '))
            if num == 1:
                for i in self.em_list:
                    if i.get_name() == name:
                        self.em_list.remove(i)
                        print(f'{name}的信息删除成功')
                    else:
                        print('没有当前员工信息！')
                break
            elif num == 2:
                ip = self.isIp()
                for i in self.em_list:
                    if i.get_ip() == ip:
                        self.em_list.remove(i)
                        print(f'{i.get_name()}的信息删除成功')
                    else:
                        print('没有当前IP信息！')
                break
            else:
                print('请输入正确的序号！')
    def modify_em(self):
        print('请输入您要修改的用户姓名')
        name = self.isName()
        for i in self.em_list:
            if i.get_name() == name:
                dept = self.isDept()
                ip = self.isIp()
                mac = self.isMac()
                i.set_dept(dept)
                i.set_ip(ip)
                i.set_mac(mac)
                print(name + '信息修改成功')
                # for else 配对，break跳出就不执行else语句。否则查询不到。就会执行else语句，注意缩进
                break
        else:
            print('没有当前用户信息')
    def query_em(self):
        print('请输入您要查询的用户姓名')
        name = self.isName()
        for i in self.em_list:
            if i.get_name() == name:
                print(i)
                break
        else:
            print('没有当前用户信息')

    # mac统一格式转换
    def transform(self, mac):
        # 首先判断是两种格式中得哪一种
        if '.' in mac:
            # 注意字符串是不可变变量
            mac = mac.replace('.', '-')
            return mac
        else:
            # 转换为列表，插入特定字符，然后转回来
            mac = mac.replace(':', '')
            ls_mac = list(mac)
            ls_mac.insert(4,'-')
            ls_mac.insert(9, '-')
            # 注意列表转字符串的方法，不然就是单个字母了
            mac = ''.join(ls_mac)
            return mac

    # 判断输入选择菜单是否合法
    def isNum(self):
        while 1:
            num = input('请输入您要选择的功能：')
            cp = re.match(r'^[1-6]$', num)
            if cp == None:
                print('请输入1-6之间的正整数！')
            else:
                return int(num)
                break

    # 分别判断名字，部门等的输入合法性
    def isName(self):
        while 1:
            name = input('姓名：')
            cp = re.match(r'^[\u4e00-\u9fa5]{2,4}$', name)
            if cp == None:
                print('请输入正确的姓名！')
            else:
                return name
                break

    def isDept(self):
        while 1:
            dept = input('部门：')
            cp = re.match(r'^[\u4e00-\u9fa5]{2,6}$', dept)
            if cp == None:
                print('请输入正确的部门！')
            else:
                return dept
                break

    def isIp(self):
        while 1:
            ip = input('IP：')
            cp = re.match(
                r'^((25[0-4]|2[0-4]\d|1\d\d|\d?\d)\.){3}(25[0-4]|2[0-4]\d|1\d\d|\d?\d)$',
                ip)
            if cp == None:
                print('请输入正确的IP格式！')
            else:
                return ip
                break

    def isMac(self):
        while 1:
            mac = input('MAC：')
            cp1 = re.match(r'^(([0-9a-fA-F]{4}\.){2}[0-9a-fA-F]{4})$',
                           mac)  # 格式1 xxxx.xxxx.xxxx
            cp2 = re.match(r'^(([0-9a-fA-F]{4}-){2}[0-9a-fA-F]{4})$',
                           mac)  # 格式2 xxxx-xxxx-xxxx
            cp3 = re.match(r'^(([0-9a-fA-F]{2}:){5}[0-9a-fA-F]{2})$',
                           mac)  #  格式3 xx:xx:xx:xx:xx:xx
            if cp1 == None and cp2 == None and cp3 == None:
                print('请输入正确的MAC格式')
            # 三种格式，下面调用转换mac函数，需要传参数进去
            elif cp2 == None:
                return self.transform(mac)
                break
            else:
                return mac
                break

    def save_em(self):
        wb = Workbook()
        ws = wb.active
        ws.title = '绑定表'
        # 表格第一行
        ws['a1'] = '姓名'
        ws['b1'] = '部门'
        ws['c1'] = 'IP'
        ws['d1'] = 'MAC'
        rowNum = len(self.em_list)
        for i in range(2,rowNum + 2):
            ws.cell(row = i,column =1).value = self.em_list[i-2].get_name()
            ws.cell(row = i,column =2).value = self.em_list[i-2].get_dept()
            ws.cell(row = i,column =3).value = self.em_list[i-2].get_ip()
            ws.cell(row = i,column =4).value = self.em_list[i-2].get_mac()
        wb.save('em_data.xlsx')

if __name__ == '__main__':
    ma = Manager()